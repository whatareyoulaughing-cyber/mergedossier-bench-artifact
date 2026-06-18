"""Build a deterministic external-audit slice from Label Studio tasks."""

from __future__ import annotations

import argparse
import importlib.util
import json
import random
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from mergedossier_bench.label_studio import export_annotation_csv_template


DEFAULT_TASKS = Path("outputs/population_ai_pr_500_20260616/reports/annotation_tasks.json")
ROOT = Path(__file__).resolve().parents[1]


def _task_id(task: dict[str, Any]) -> str:
    return str(task.get("data", {}).get("instance_id", ""))


def select_tasks(tasks: list[dict[str, Any]], n: int, seed: int) -> list[dict[str, Any]]:
    primary = [task for task in tasks if not task.get("data", {}).get("is_reliability_repeat")]
    unique = {_task_id(task): task for task in primary if _task_id(task)}
    ordered = [unique[key] for key in sorted(unique)]
    if n > len(ordered):
        raise ValueError(f"Requested {n} tasks, but only {len(ordered)} unique primary tasks are available.")
    rng = random.Random(seed)
    selected = rng.sample(ordered, n)
    return sorted(selected, key=_task_id)


def write_instructions(out_dir: Path, n: int, seed: int) -> None:
    lines = [
        "# External Audit Slice Instructions",
        "",
        "This packet contains a deterministic 10% external audit slice for the AIDev-pop handoff-evidence study.",
        "It is designed for a second operator or external reviewer to independently code the same evidence-availability families.",
        "",
        f"- Tasks: {n}",
        f"- Sampling seed: {seed}",
        "- Source task file: `outputs/population_ai_pr_500_20260616/reports/annotation_tasks.json`",
        "",
        "## Files",
        "",
        "- `external_audit_tasks.json`: Label Studio import tasks.",
        "- `external_audit_sheet.csv`: spreadsheet-friendly audit sheet with blank audit-code columns.",
        "- `external_audit_sheet.xlsx`: Excel workbook with dropdowns, instructions, and completion checks when openpyxl is available.",
        "- `external_audit_manifest.json`: selected instance IDs and packet metadata.",
        "",
        "## Coding Boundary",
        "",
        "Use the existing audit-code vocabulary: `present`, `partially_present`, `missing`, and `not_applicable`.",
        "The task is to assess visible review-evidence availability, not patch correctness, mergeability, reviewer utility, or AI-vs-human effects.",
        "If completed, these rows can support an external audit slice. Until completed, the main paper should continue to report single-operator self-consistency only.",
        "",
        "## After Completion",
        "",
        "Save the completed sheet as a CSV and run:",
        "",
        "```bash",
        "python scripts/analyze_external_audit_slice.py \\",
        "  --primary outputs/population_ai_pr_500_20260616/reports/annotation_sheet_completed.csv \\",
        "  --external <completed_external_audit_sheet.csv> \\",
        "  --manifest outputs/external_audit_slice_20260617/external_audit_manifest.json \\",
        "  --out outputs/external_audit_analysis_20260617",
        "```",
        "",
        "The analysis emits `external_audit_summary.json/md`,",
        "`external_audit_agreement_by_category.csv`, a LaTeX-ready external-audit table,",
        "and disagreement files. If any audit-code cell is blank or invalid, the report",
        "stays `incomplete` and must not be cited as external agreement.",
        "",
    ]
    (out_dir / "README_external_audit.md").write_text("\n".join(lines), encoding="utf-8")


def _load_workbook_builder() -> Any:
    script_path = ROOT / "scripts" / "export_annotation_workbook.py"
    spec = importlib.util.spec_from_file_location("export_annotation_workbook", script_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load workbook exporter: {script_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module.build_workbook


def build_external_audit_workbook(csv_path: Path, workbook_path: Path) -> tuple[str | None, str]:
    try:
        build_workbook = _load_workbook_builder()
        out = build_workbook(csv_path, workbook_path)
        try:
            from openpyxl import load_workbook

            wb = load_workbook(out)
            instructions = wb["Instructions"]
            instructions["A1"] = "MergeDossier-Bench External Audit Slice Workbook"
            instructions["A3"] = "Fill every *_label cell using the dropdown values."
            instructions["A4"] = "Code these 50 tasks independently from the primary operator. Do not look at the completed primary audit sheet while coding."
            instructions["A5"] = "Do not judge code correctness or mergeability. Code only visible review-evidence availability in the dossier."
            instructions["C7"] = "After completion, save/export the Annotation sheet as CSV, then run:"
            instructions["C8"] = "python scripts/analyze_external_audit_slice.py --external <completed_external_audit_sheet.csv>"
            instructions["C10"] = "Expected analysis outputs:"
            instructions["C11"] = "outputs/external_audit_analysis_20260617/external_audit_summary.json"
            instructions["C12"] = "outputs/external_audit_analysis_20260617/external_audit_agreement_by_category.csv"
            wb.save(out)
        except Exception:
            # Workbook creation succeeded; instruction customization is best-effort.
            pass
        return str(out), "created"
    except SystemExit as exc:
        return None, f"skipped: {exc}"
    except ModuleNotFoundError as exc:
        return None, f"skipped: missing dependency {exc.name}"


def build_external_audit_slice(
    tasks_path: str | Path,
    out_dir: str | Path,
    n: int = 50,
    seed: int = 20260617,
    make_workbook: bool = True,
) -> dict[str, Any]:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    tasks = json.loads(Path(tasks_path).read_text(encoding="utf-8"))
    selected = select_tasks(tasks, n=n, seed=seed)
    tasks_out = out / "external_audit_tasks.json"
    tasks_out.write_text(json.dumps(selected, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    csv_out = out / "external_audit_sheet.csv"
    rows = export_annotation_csv_template(tasks_out, csv_out, annotator_id="external_auditor")
    workbook_out: str | None = None
    workbook_status = "not_requested"
    if make_workbook:
        workbook_out, workbook_status = build_external_audit_workbook(csv_out, out / "external_audit_sheet.xlsx")
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_tasks": str(tasks_path),
        "tasks_available": len(tasks),
        "selected_tasks": len(selected),
        "seed": seed,
        "selected_instance_ids": [_task_id(task) for task in selected],
        "tasks_json": str(tasks_out),
        "audit_csv": str(csv_out),
        "audit_workbook": workbook_out,
        "audit_workbook_status": workbook_status,
        "rows": len(rows),
        "claim_boundary": (
            "External audit slice packet only. Do not claim inter-rater reliability unless the slice is completed "
            "by an independent operator and analyzed separately."
        ),
    }
    (out / "external_audit_manifest.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    write_instructions(out, n=len(selected), seed=seed)
    return summary


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build deterministic external audit slice")
    parser.add_argument("--tasks", default=str(DEFAULT_TASKS), help="Source Label Studio task JSON")
    parser.add_argument("--out", default="outputs/external_audit_slice_20260617", help="Output directory")
    parser.add_argument("--n", type=int, default=50, help="Number of primary tasks to select")
    parser.add_argument("--seed", type=int, default=20260617, help="Deterministic random seed")
    parser.add_argument("--no-workbook", action="store_true", help="Do not attempt to create external_audit_sheet.xlsx")
    args = parser.parse_args(argv)
    summary = build_external_audit_slice(args.tasks, args.out, n=args.n, seed=args.seed, make_workbook=not args.no_workbook)
    print(f"External audit slice written: {args.out} ({summary['selected_tasks']} tasks)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

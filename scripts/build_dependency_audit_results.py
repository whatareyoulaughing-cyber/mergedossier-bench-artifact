"""Build focused dependency-evidence audit results from a completed sheet."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from mergedossier_bench.population import wilson_interval


LABEL_VALUES = {"present", "partially_present", "missing", "not_applicable"}


def read_sheet(path: str | Path, sheet_name: str | None = None) -> list[dict[str, str]]:
    source = Path(path)
    if source.suffix.lower() == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as f:
            return [{key: value or "" for key, value in row.items()} for row in csv.DictReader(f)]
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise SystemExit("openpyxl is required to read dependency audit workbooks.") from exc
    wb = load_workbook(source, read_only=True, data_only=True)
    selected = sheet_name or ("DependencyAudit" if "DependencyAudit" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[selected]
    rows = ws.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    output: list[dict[str, str]] = []
    for values in rows:
        record = {headers[index]: str(value or "").strip() for index, value in enumerate(values)}
        if record.get("instance_id"):
            output.append(record)
    return output


def build_dependency_audit_summary(rows: list[dict[str, str]], sample_size: int = 500) -> dict[str, Any]:
    errors: list[str] = []
    label_counts: Counter[str] = Counter()
    for index, row in enumerate(rows, start=2):
        label = str(row.get("dependency_evidence_label", "")).strip()
        if not label:
            errors.append(f"Row {index}: missing dependency_evidence_label")
            continue
        if label not in LABEL_VALUES:
            errors.append(f"Row {index}: invalid dependency_evidence_label {label!r}")
            continue
        label_counts[label] += 1
    candidate_count = len(rows)
    applicable = candidate_count - label_counts["not_applicable"]
    positive = label_counts["present"] + label_counts["partially_present"]
    missing = label_counts["missing"]
    interval = wilson_interval(positive, applicable)
    return {
        "valid": not errors,
        "errors": errors,
        "sample_size": sample_size,
        "dependency_sensitive_candidates": candidate_count,
        "candidate_rate": round(candidate_count / sample_size, 4) if sample_size else None,
        "applicable_candidates": applicable,
        "positive_candidates": positive,
        "missing_candidates": missing,
        "label_counts": {label: label_counts.get(label, 0) for label in ("present", "partially_present", "missing", "not_applicable")},
        "coverage_rate": interval["proportion"],
        "coverage_ci_lower": interval["lower"],
        "coverage_ci_upper": interval["upper"],
        "missing_rate": round(missing / applicable, 4) if applicable else None,
        "interpretation": (
            "Dependency estimates apply only to the dependency-sensitive candidate subset detected by manifest/lockfile changes. "
            "They are a secondary focused audit, not a full-sample dependency denominator."
        ),
    }


def _pct(value: Any) -> str:
    if value is None or value == "":
        return "--"
    return f"{float(value) * 100:.1f}%"


def write_results(summary: dict[str, Any], out_dir: str | Path) -> None:
    output = Path(out_dir)
    output.mkdir(parents=True, exist_ok=True)
    (output / "dependency_audit_results.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    with (output / "dependency_audit_results.csv").open("w", encoding="utf-8", newline="") as f:
        fields = [
            "dependency_sensitive_candidates",
            "candidate_rate",
            "applicable_candidates",
            "positive_candidates",
            "coverage_rate",
            "coverage_ci_lower",
            "coverage_ci_upper",
            "missing_candidates",
            "missing_rate",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerow({field: summary.get(field, "") for field in fields})
    md = [
        "# Dependency-sensitive Audit Results",
        "",
        summary["interpretation"],
        "",
        f"- Dependency-sensitive candidates: {summary['dependency_sensitive_candidates']} / {summary['sample_size']} ({_pct(summary['candidate_rate'])})",
        f"- Applicable candidates: {summary['applicable_candidates']}",
        f"- Positive dependency evidence: {summary['positive_candidates']} ({_pct(summary['coverage_rate'])}; Wilson 95% CI {_pct(summary['coverage_ci_lower'])}--{_pct(summary['coverage_ci_upper'])})",
        f"- Missing dependency evidence: {summary['missing_candidates']} ({_pct(summary['missing_rate'])})",
        f"- Label counts: {summary['label_counts']}",
        "",
    ]
    (output / "dependency_audit_results.md").write_text("\n".join(md), encoding="utf-8")
    tex = [
        r"\begin{table}[t]",
        r"\centering",
        r"\caption{Focused dependency-sensitive audit results.}",
        r"\label{tab:dependency-sensitive-audit}",
        r"\footnotesize",
        r"\begin{tabular}{@{}lr@{}}",
        r"\toprule",
        r"Metric & Value \\",
        r"\midrule",
        f"Candidate PRs & {summary['dependency_sensitive_candidates']} / {summary['sample_size']} ({_pct(summary['candidate_rate'])}) \\\\",
        f"Positive evidence among candidates & {summary['positive_candidates']} / {summary['applicable_candidates']} ({_pct(summary['coverage_rate'])}) \\\\",
        f"Wilson 95\\% CI & {_pct(summary['coverage_ci_lower'])}--{_pct(summary['coverage_ci_upper'])} \\\\",
        f"Missing among candidates & {summary['missing_candidates']} ({_pct(summary['missing_rate'])}) \\\\",
        r"\bottomrule",
        r"\end{tabular}",
        r"\end{table}",
        "",
    ]
    (output / "dependency_audit_table.tex").write_text("\n".join(tex), encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build dependency-sensitive audit result tables")
    parser.add_argument("--annotations", required=True, help="Completed dependency audit CSV or XLSX")
    parser.add_argument("--out", required=True, help="Output directory")
    parser.add_argument("--sheet", help="Workbook sheet name; defaults to DependencyAudit")
    parser.add_argument("--sample-size", type=int, default=500)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    rows = read_sheet(args.annotations, sheet_name=args.sheet)
    summary = build_dependency_audit_summary(rows, sample_size=args.sample_size)
    write_results(summary, args.out)
    if not summary["valid"]:
        print(f"Dependency audit has {len(summary['errors'])} error(s); see dependency_audit_results.json")
        return 1
    print(
        "Dependency audit results written: "
        f"{summary['positive_candidates']} positive of {summary['applicable_candidates']} applicable candidates -> {args.out}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

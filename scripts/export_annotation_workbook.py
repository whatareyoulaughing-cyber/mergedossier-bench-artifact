"""Create an Excel workbook for single-operator MergeDossier audit coding."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any


LABEL_VALUES = ("present", "partially_present", "missing", "not_applicable")
LABEL_SUFFIX = "_label"
COMMENT_SUFFIX = "_comment"


def label_columns(headers: list[str]) -> list[str]:
    return [header for header in headers if header.endswith(LABEL_SUFFIX)]


def comment_columns(headers: list[str]) -> list[str]:
    return [header for header in headers if header.endswith(COMMENT_SUFFIX)]


def load_csv(path: str | Path) -> tuple[list[str], list[dict[str, str]]]:
    csv_path = Path(path)
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    if not headers:
        raise ValueError(f"CSV has no header row: {csv_path}")
    if not rows:
        raise ValueError(f"CSV has no annotation rows: {csv_path}")
    return headers, rows


def _require_openpyxl() -> Any:
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required to build the Excel workbook. "
            "Use the Codex bundled Python runtime or install openpyxl."
        ) from exc
    return {
        "openpyxl": openpyxl,
        "Workbook": Workbook,
        "Comment": Comment,
        "FormulaRule": FormulaRule,
        "Alignment": Alignment,
        "Border": Border,
        "Font": Font,
        "PatternFill": PatternFill,
        "Side": Side,
        "DataValidation": DataValidation,
    }


def build_workbook(csv_path: str | Path, out: str | Path) -> Path:
    api = _require_openpyxl()
    headers, rows = load_csv(csv_path)
    label_cols = label_columns(headers)
    comment_cols = comment_columns(headers)

    Workbook = api["Workbook"]
    Comment = api["Comment"]
    FormulaRule = api["FormulaRule"]
    Alignment = api["Alignment"]
    Border = api["Border"]
    Font = api["Font"]
    PatternFill = api["PatternFill"]
    Side = api["Side"]
    DataValidation = api["DataValidation"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Annotation"
    instructions = wb.create_sheet("Instructions")
    values = wb.create_sheet("LabelValues")
    values.sheet_state = "hidden"

    instructions["A1"] = "MergeDossier-Bench Review-Evidence Availability Audit Workbook"
    instructions["A1"].font = Font(bold=True, size=16, color="1F2937")
    instructions["A3"] = "Fill every *_label cell using the dropdown values."
    instructions["A4"] = "Complete non-repeat rows first. If possible, wait about 48 hours, then complete rows where is_reliability_repeat is TRUE."
    instructions["A5"] = "Do not judge code correctness or mergeability. Code only visible review-evidence availability in the dossier."
    instructions["A7"] = "Allowed labels"
    instructions["A7"].font = Font(bold=True)
    for idx, value in enumerate(LABEL_VALUES, start=8):
        instructions[f"A{idx}"] = value
    instructions["C7"] = "Save completed workbook and export/save the Annotation sheet as CSV:"
    instructions["C8"] = "outputs/real_pilot_mixed_source_annotation_sheet_completed_20260613.csv"
    instructions["C10"] = "Then run:"
    instructions["C11"] = "python -m mergedossier_bench.cli validate-annotation-csv --annotations outputs/real_pilot_mixed_source_annotation_sheet_completed_20260613.csv"
    instructions["C12"] = "python -m mergedossier_bench.cli annotation-stats --annotations outputs/real_pilot_mixed_source_annotation_sheet_completed_20260613.csv --out outputs/real_pilot_mixed_source_annotation_stats_20260613"
    for column in ("A", "C"):
        instructions.column_dimensions[column].width = 72
    for row in instructions.iter_rows(min_row=1, max_row=12, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    values["A1"] = "label"
    for idx, value in enumerate(LABEL_VALUES, start=2):
        values[f"A{idx}"] = value
    validation_formula = f"=LabelValues!$A$2:$A${len(LABEL_VALUES) + 1}"

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    label_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    comment_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    repeat_fill = PatternFill(fill_type="solid", fgColor="EADCF8")
    missing_fill = PatternFill(fill_type="solid", fgColor="FCE4D6")
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border
    for idx, header in enumerate(headers, start=1):
        column_letter = ws.cell(row=1, column=idx).column_letter
        if header in label_cols:
            ws.column_dimensions[column_letter].width = 22
            for cell in ws[column_letter][1:]:
                cell.fill = label_fill
        elif header in comment_cols:
            ws.column_dimensions[column_letter].width = 34
            for cell in ws[column_letter][1:]:
                cell.fill = comment_fill
        elif header == "dossier_text":
            ws.column_dimensions[column_letter].width = 72
        elif header in {"title", "missing_evidence", "pr_url", "source"}:
            ws.column_dimensions[column_letter].width = 36
        else:
            ws.column_dimensions[column_letter].width = 18
    ws.row_dimensions[1].height = 42
    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_num in range(2, len(rows) + 2):
        ws.row_dimensions[row_num].height = 58

    dv = DataValidation(
        type="list",
        formula1=validation_formula,
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid label",
        error="Use present, partially_present, missing, or not_applicable.",
        promptTitle="Evidence label",
        prompt="Choose exactly one allowed label.",
    )
    ws.add_data_validation(dv)
    for header in label_cols:
        col = headers.index(header) + 1
        letter = ws.cell(row=1, column=col).column_letter
        dv.add(f"{letter}2:{letter}{len(rows) + 1}")
        ws.cell(row=1, column=col).comment = Comment(
            "Required. Use the dropdown only: present, partially_present, missing, not_applicable.",
            "Codex",
        )

    if "is_reliability_repeat" in headers:
        repeat_col = headers.index("is_reliability_repeat") + 1
        repeat_letter = ws.cell(row=1, column=repeat_col).column_letter
        ws.conditional_formatting.add(
            f"A2:{ws.cell(row=1, column=len(headers)).column_letter}{len(rows) + 1}",
            FormulaRule(formula=[f'=${repeat_letter}2=TRUE'], fill=repeat_fill),
        )

    for header in label_cols:
        col = headers.index(header) + 1
        letter = ws.cell(row=1, column=col).column_letter
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{len(rows) + 1}",
            FormulaRule(formula=[f'ISBLANK({letter}2)'], fill=missing_fill),
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(headers)).column_letter}{len(rows) + 1}"

    completion = wb.create_sheet("Completion")
    completion["A1"] = "Completion Check"
    completion["A1"].font = Font(bold=True, size=16)
    completion["A3"] = "Rows"
    completion["B3"] = len(rows)
    completion["A4"] = "Required label cells"
    completion["B4"] = len(rows) * len(label_cols)
    completion["A5"] = "Filled label cells"
    count_terms = []
    for header in label_cols:
        col = headers.index(header) + 1
        letter = ws.cell(row=1, column=col).column_letter
        count_terms.append(f"COUNTA(Annotation!{letter}2:{letter}{len(rows) + 1})")
    completion["B5"] = "=" + "+".join(count_terms)
    completion["A6"] = "Ready for validation?"
    completion["B6"] = '=IF(B5=B4,"YES","NO")'
    completion["A8"] = "Note"
    completion["B8"] = "This sheet is a convenience check. The authoritative check is validate-annotation-csv."
    completion.column_dimensions["A"].width = 28
    completion.column_dimensions["B"].width = 80
    for row in completion.iter_rows(min_row=1, max_row=8, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export spreadsheet-friendly annotation workbook")
    parser.add_argument("--csv", required=True, help="Input annotation CSV template")
    parser.add_argument("--out", required=True, help="Output .xlsx workbook")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    out = build_workbook(args.csv, args.out)
    print(f"Annotation workbook written: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""Build a focused dependency-evidence audit sheet from a population manifest."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from mergedossier_bench.population import read_csv_rows


DEPENDENCY_FILE_PATTERNS: tuple[str, ...] = (
    r"(^|/)package\.json$",
    r"(^|/)package-lock\.json$",
    r"(^|/)pnpm-lock\.yaml$",
    r"(^|/)yarn\.lock$",
    r"(^|/)requirements\.txt$",
    r"(^|/)pyproject\.toml$",
    r"(^|/)poetry\.lock$",
    r"(^|/)go\.mod$",
    r"(^|/)go\.sum$",
    r"(^|/)Cargo\.toml$",
    r"(^|/)Cargo\.lock$",
    r"(^|/)pom\.xml$",
    r"(^|/)build\.gradle$",
    r"(^|/)build\.gradle\.kts$",
    r"(^|/)Gemfile$",
    r"(^|/)Gemfile\.lock$",
    r"(^|/)composer\.json$",
    r"(^|/)composer\.lock$",
    r"(^|/)Package\.swift$",
    r"(^|/)mix\.exs$",
    r"(^|/)mix\.lock$",
    r"(^|/)pubspec\.yaml$",
    r"(^|/)bun\.lockb$",
)

DEPENDENCY_RE = re.compile("|".join(f"(?:{pattern})" for pattern in DEPENDENCY_FILE_PATTERNS), re.IGNORECASE)

OUTPUT_COLUMNS: tuple[str, ...] = (
    "annotator_id",
    "instance_id",
    "repo",
    "pr_number",
    "pr_url",
    "title",
    "agent_name",
    "language",
    "outcome",
    "dependency_files",
    "files_changed",
    "body",
    "commit_messages",
    "comments",
    "reviews",
    "dependency_evidence_label",
    "dependency_evidence_comment",
)


def split_file_list(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[|;\n,]", value or "") if part.strip()]


def dependency_files(files_changed: str) -> list[str]:
    return [filename for filename in split_file_list(files_changed) if DEPENDENCY_RE.search(filename)]


def truncate(value: Any, limit: int = 1200) -> str:
    text = str(value or "").replace("\r\n", "\n").strip()
    return text[:limit]


def build_dependency_audit_rows(manifest_rows: list[dict[str, str]], annotator_id: str = "solo") -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in manifest_rows:
        deps = dependency_files(row.get("files_changed", ""))
        if not deps:
            continue
        rows.append(
            {
                "annotator_id": annotator_id,
                "instance_id": row.get("instance_id", ""),
                "repo": row.get("repo", ""),
                "pr_number": row.get("pr_number", ""),
                "pr_url": row.get("pr_url", ""),
                "title": row.get("title", ""),
                "agent_name": row.get("agent_name", ""),
                "language": row.get("language", ""),
                "outcome": row.get("outcome", ""),
                "dependency_files": " | ".join(deps),
                "files_changed": truncate(row.get("files_changed", ""), 2000),
                "body": truncate(row.get("body", "")),
                "commit_messages": truncate(row.get("commit_messages", "")),
                "comments": truncate(row.get("comments", "")),
                "reviews": truncate(row.get("reviews", "")),
                "dependency_evidence_label": "",
                "dependency_evidence_comment": "",
            }
        )
    return rows


def write_csv(path: str | Path, rows: list[dict[str, str]]) -> None:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(csv_path: str | Path, out_path: str | Path) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError:
        return False

    csv_rows = list(csv.DictReader(Path(csv_path).open("r", encoding="utf-8-sig", newline="")))
    wb = Workbook()
    ws = wb.active
    ws.title = "DependencyAudit"
    ws.append(list(OUTPUT_COLUMNS))
    for row in csv_rows:
        ws.append([row.get(column, "") for column in OUTPUT_COLUMNS])
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for idx, column in enumerate(OUTPUT_COLUMNS, start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = 24 if column not in {"body", "files_changed", "commit_messages", "comments", "reviews"} else 56
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    validation = DataValidation(
        type="list",
        formula1='"present,partially_present,missing,not_applicable"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid label",
        error="Use present, partially_present, missing, or not_applicable.",
    )
    ws.add_data_validation(validation)
    label_col = OUTPUT_COLUMNS.index("dependency_evidence_label") + 1
    label_letter = ws.cell(row=1, column=label_col).column_letter
    validation.add(f"{label_letter}2:{label_letter}{len(csv_rows) + 1}")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(OUTPUT_COLUMNS)).column_letter}{len(csv_rows) + 1}"
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return True


def write_summary(out_dir: str | Path, manifest_rows: list[dict[str, str]], audit_rows: list[dict[str, str]], workbook_written: bool) -> dict[str, Any]:
    dep_counter = Counter()
    for row in audit_rows:
        for filename in split_file_list(row.get("dependency_files", "")):
            dep_counter[Path(filename).name] += 1
    summary = {
        "manifest_rows": len(manifest_rows),
        "dependency_sensitive_candidates": len(audit_rows),
        "candidate_rate": round(len(audit_rows) / len(manifest_rows), 4) if manifest_rows else 0,
        "top_dependency_files": dict(dep_counter.most_common(20)),
        "workbook_written": workbook_written,
        "interpretation": (
            "This packet identifies PRs with dependency manifest or lockfile changes for focused follow-up. "
            "It is not a completed dependency-evidence coverage estimate until the dependency_evidence_label column is annotated."
        ),
    }
    output = Path(out_dir)
    output.mkdir(parents=True, exist_ok=True)
    (output / "dependency_candidate_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    (output / "README_ZH.md").write_text(
        "\n".join(
            [
                "# Dependency-sensitive audit 下一步",
                "",
                "这个包只用于补 Dependency evidence 的 targeted follow-up，不需要重标全部 evidence families。",
                "",
                f"- 500 sample rows: {summary['manifest_rows']}",
                f"- Dependency-sensitive candidates: {summary['dependency_sensitive_candidates']}",
                "",
                "请填写 `dependency_audit_sheet.xlsx` 或 `dependency_audit_sheet.csv` 的 `dependency_evidence_label` 列。",
                "允许值：`present`, `partially_present`, `missing`, `not_applicable`。",
                "",
                "完成前，论文只能说 Dependency 被主标注保守排除，不能报告 Dependency coverage。",
                "",
            ]
        ),
        encoding="utf-8",
    )
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build dependency-sensitive follow-up audit sheet")
    parser.add_argument("--manifest", required=True, help="Population sample manifest CSV")
    parser.add_argument("--out", required=True, help="Output directory")
    parser.add_argument("--annotator-id", default="solo")
    parser.add_argument("--no-workbook", action="store_true")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    out_dir = Path(args.out)
    manifest_rows = read_csv_rows(args.manifest)
    audit_rows = build_dependency_audit_rows(manifest_rows, annotator_id=args.annotator_id)
    csv_path = out_dir / "dependency_audit_sheet.csv"
    write_csv(csv_path, audit_rows)
    workbook_written = False
    if not args.no_workbook:
        workbook_written = write_workbook(csv_path, out_dir / "dependency_audit_sheet.xlsx")
    summary = write_summary(out_dir, manifest_rows, audit_rows, workbook_written)
    print(
        "Dependency audit sheet written: "
        f"{summary['dependency_sensitive_candidates']} candidates of {summary['manifest_rows']} rows -> {out_dir}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

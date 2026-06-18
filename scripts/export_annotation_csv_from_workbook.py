"""Export the Annotation sheet from an Excel workbook to CSV."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required to export the workbook. "
            "Use the Codex bundled Python runtime or install with: pip install -e \".[workbook]\""
        ) from exc
    return openpyxl


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def export_annotation_csv(
    workbook_path: str | Path,
    out: str | Path,
    sheet_name: str = "Annotation",
) -> Path:
    """Write the annotation worksheet as a UTF-8-SIG CSV file."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook does not contain a {sheet_name!r} sheet: {workbook_path}")

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet {sheet_name!r} is empty: {workbook_path}")
    headers = [_cell_text(value).strip() for value in rows[0]]
    if not any(headers):
        raise ValueError(f"Sheet {sheet_name!r} has no header row: {workbook_path}")
    if "instance_id" not in headers:
        raise ValueError(f"Sheet {sheet_name!r} does not look like an annotation sheet: missing instance_id")

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows[1:]:
            padded = list(row) + [None] * (len(headers) - len(row))
            writer.writerow([_cell_text(value) for value in padded[: len(headers)]])
    return out_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export workbook Annotation sheet to CSV")
    parser.add_argument("--workbook", required=True, help="Input .xlsx annotation workbook")
    parser.add_argument("--out", required=True, help="Output completed annotation CSV")
    parser.add_argument("--sheet", default="Annotation", help="Workbook sheet name to export")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    out = export_annotation_csv(args.workbook, args.out, sheet_name=args.sheet)
    print(f"Annotation CSV exported: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
